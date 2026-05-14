import streamlit as st
import pandas as pd
import sqlite3
import altair as alt
from io import BytesIO
from datetime import datetime

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows

from sigem_db import get_db_path


# =====================================================
# OBTENER INCIDENCIAS
# =====================================================

def obtener_incidencias():

    conn = sqlite3.connect(get_db_path("logistica"))

    query = """
        SELECT
            folio_incidencia,
            fecha,
            modulo,
            proceso,
            tipo_incidencia,
            prioridad,
            estatus,
            folio_referencia,
            codigo_transporte,
            folio_hoja_carga,
            pedido,
            codigo_material,
            descripcion,
            cantidad,
            cliente,
            destino,
            bodega,
            ubicacion,
            transportista,
            vehiculo,
            placas,
            operador,
            responsable,
            descripcion_incidencia,
            causa,
            solucion,
            fecha_solucion,
            usuario_registro,
            usuario_cierre,
            observaciones,
            fecha_creacion
        FROM incidencias
        WHERE
            codigo_transporte IS NOT NULL
            AND TRIM(codigo_transporte) <> ''
            AND UPPER(TRIM(estatus)) NOT IN (
                'CONCLUIDA',
                'CONCLUIDO',
                'FINALIZADA',
                'FINALIZADO',
                'CERRADA',
                'CERRADO',
                'RESUELTA',
                'RESUELTO',
                'CANCELADA',
                'CANCELADO'
            )
        ORDER BY fecha DESC, folio_incidencia DESC
    """

    try:
        df = pd.read_sql_query(query, conn)

    except Exception:

        query = """
            SELECT
                folio_incidencia,
                fecha,
                modulo,
                proceso,
                tipo_incidencia,
                prioridad,
                estatus,
                folio_referencia,
                '' AS codigo_transporte,
                folio_hoja_carga,
                pedido,
                codigo_material,
                descripcion,
                cantidad,
                cliente,
                destino,
                bodega,
                ubicacion,
                transportista,
                vehiculo,
                placas,
                operador,
                responsable,
                descripcion_incidencia,
                causa,
                solucion,
                fecha_solucion,
                usuario_registro,
                usuario_cierre,
                observaciones,
                fecha_creacion
            FROM incidencias
            WHERE
                UPPER(TRIM(estatus)) NOT IN (
                    'CONCLUIDA',
                    'CONCLUIDO',
                    'FINALIZADA',
                    'FINALIZADO',
                    'CERRADA',
                    'CERRADO',
                    'RESUELTA',
                    'RESUELTO',
                    'CANCELADA',
                    'CANCELADO'
                )
            ORDER BY fecha DESC, folio_incidencia DESC
        """

        df = pd.read_sql_query(query, conn)

    conn.close()

    return df


# =====================================================
# PREPARAR DATOS
# =====================================================

def preparar_incidencias(df):

    df = df.copy()
    df = df.fillna("")

    if "codigo_transporte" not in df.columns:
        df["codigo_transporte"] = ""

    df["transporte_mostrar"] = (
        df["codigo_transporte"]
        .astype(str)
        .str.strip()
    )

    df = df[
        df["transporte_mostrar"] != ""
    ]

    return df


# =====================================================
# EXPORTAR EXCEL
# =====================================================

def generar_excel_incidencias(df_filtrado, total, abiertas, proceso, criticas):

    wb = Workbook()
    ws = wb.active
    ws.title = "Dashboard"

    azul = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
    font_blanca = Font(color="FFFFFF", bold=True)
    borde = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin")
    )

    ws["A1"] = "SIGEM - Dashboard incidencias por transporte"
    ws["A1"].font = Font(bold=True, size=16)

    indicadores = [
        ("Total incidencias", total),
        ("Abiertas", abiertas),
        ("En proceso", proceso),
        ("Críticas", criticas)
    ]

    fila = 3
    for titulo, valor in indicadores:
        ws[f"A{fila}"] = titulo
        ws[f"B{fila}"] = valor
        ws[f"A{fila}"].fill = azul
        ws[f"A{fila}"].font = font_blanca
        ws[f"A{fila}"].border = borde
        fila += 1

    ws2 = wb.create_sheet(title="Detalle incidencias")

    for r in dataframe_to_rows(df_filtrado, index=False, header=True):
        ws2.append(r)

    for cell in ws2[1]:
        cell.fill = azul
        cell.font = font_blanca
        cell.border = borde

    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    return buffer


# =====================================================
# GRAFICA MATRIZ
# =====================================================

def pintar_matriz_incidencias(df):

    orden_estatus = [
        "Abierta",
        "En proceso",
        "En seguimiento"
    ]

    mapa_colores = {
        "Baja": "#6B7280",
        "Media": "#EAB308",
        "Alta": "#F97316",
        "Crítica": "#DC2626"
    }

    df_grafica = df.copy()

    df_grafica["estatus"] = (
        df_grafica["estatus"]
        .fillna("Abierta")
        .astype(str)
        .str.strip()
    )

    df_grafica["prioridad"] = (
        df_grafica["prioridad"]
        .fillna("Media")
        .astype(str)
        .str.strip()
    )

    df_grafica["incidencia"] = (
        df_grafica["folio_incidencia"]
        .fillna("")
        .astype(str)
        .str.strip()
    )

    ancho_grafica = max(
        1300,
        len(df_grafica) * 95
    )

    chart = (
        alt.Chart(df_grafica)
        .mark_rect(
            cornerRadius=5,
            width=62,
            height=26
        )
        .encode(
            x=alt.X(
                "incidencia:N",
                title="Folio incidencia",
                sort=None,
                axis=alt.Axis(
                    labelAngle=-45,
                    labelLimit=130,
                    grid=True,
                    tickSize=0
                )
            ),
            y=alt.Y(
                "estatus:N",
                sort=orden_estatus,
                title="Estatus incidencia",
                axis=alt.Axis(
                    grid=True,
                    tickSize=0
                )
            ),
            color=alt.Color(
                "prioridad:N",
                scale=alt.Scale(
                    domain=list(mapa_colores.keys()),
                    range=list(mapa_colores.values())
                ),
                legend=alt.Legend(title="Prioridad")
            ),
            tooltip=[
                alt.Tooltip("folio_incidencia:N", title="Folio incidencia"),
                alt.Tooltip("transporte_mostrar:N", title="Transporte"),
                alt.Tooltip("estatus:N", title="Estatus incidencia"),
                alt.Tooltip("prioridad:N", title="Prioridad"),
                alt.Tooltip("tipo_incidencia:N", title="Tipo incidencia"),
                alt.Tooltip("cliente:N", title="Cliente"),
                alt.Tooltip("destino:N", title="Destino"),
                alt.Tooltip("transportista:N", title="Transportista"),
                alt.Tooltip("vehiculo:N", title="Vehículo"),
                alt.Tooltip("placas:N", title="Placas"),
                alt.Tooltip("operador:N", title="Operador"),
                alt.Tooltip("responsable:N", title="Responsable"),
                alt.Tooltip("descripcion_incidencia:N", title="Descripción"),
                alt.Tooltip("fecha:N", title="Fecha incidencia")
            ]
        )
        .properties(
            width=ancho_grafica,
            height=520
        )
        .configure_axis(
            grid=True,
            gridColor="#D1D5DB",
            gridDash=[3, 2],
            domain=False,
            labelColor="#374151",
            titleColor="#111827"
        )
        .configure_view(
            stroke="#D1D5DB"
        )
    )

    st.altair_chart(
        chart,
        use_container_width=False
    )


# =====================================================
# APP
# =====================================================

def dashboard_incidencias_app():

    st.title("📊 Dashboard incidencias por transporte")

    try:
        df = obtener_incidencias()

    except Exception as e:
        st.error("❌ Error consultando incidencias")
        st.exception(e)
        return

    if df.empty:
        st.warning("No existen incidencias activas por transporte.")
        return

    df = preparar_incidencias(df)

    if df.empty:
        st.warning("No existen incidencias activas con transporte asignado.")
        return

    st.subheader("🔎 Filtros")

    col1, col2, col3, col4 = st.columns(4)

    with col1:
        filtro_transporte = st.text_input("Transporte")

    with col2:
        filtro_transportista = st.text_input("Transportista")

    with col3:
        filtro_estatus = st.selectbox(
            "Estatus",
            ["Todos"] + sorted(df["estatus"].astype(str).unique().tolist())
        )

    with col4:
        filtro_prioridad = st.selectbox(
            "Prioridad",
            ["Todos"] + sorted(df["prioridad"].astype(str).unique().tolist())
        )

    col5, col6 = st.columns(2)

    with col5:
        filtro_cliente = st.text_input("Cliente")

    with col6:
        filtro_operador = st.text_input("Operador")

    df_filtrado = df.copy()

    if filtro_transporte:
        df_filtrado = df_filtrado[
            df_filtrado["transporte_mostrar"]
            .astype(str)
            .str.contains(filtro_transporte, case=False, na=False)
        ]

    if filtro_transportista:
        df_filtrado = df_filtrado[
            df_filtrado["transportista"]
            .astype(str)
            .str.contains(filtro_transportista, case=False, na=False)
        ]

    if filtro_cliente:
        df_filtrado = df_filtrado[
            df_filtrado["cliente"]
            .astype(str)
            .str.contains(filtro_cliente, case=False, na=False)
        ]

    if filtro_operador:
        df_filtrado = df_filtrado[
            df_filtrado["operador"]
            .astype(str)
            .str.contains(filtro_operador, case=False, na=False)
        ]

    if filtro_estatus != "Todos":
        df_filtrado = df_filtrado[
            df_filtrado["estatus"].astype(str) == filtro_estatus
        ]

    if filtro_prioridad != "Todos":
        df_filtrado = df_filtrado[
            df_filtrado["prioridad"].astype(str) == filtro_prioridad
        ]

    total = len(df_filtrado)

    abiertas = df_filtrado[
        df_filtrado["estatus"]
        .astype(str)
        .str.contains("Abierta", case=False, na=False)
    ].shape[0]

    proceso = df_filtrado[
        df_filtrado["estatus"]
        .astype(str)
        .str.contains("proceso|seguimiento", case=False, na=False)
    ].shape[0]

    criticas = df_filtrado[
        df_filtrado["prioridad"]
        .astype(str)
        .str.contains("Crítica|Critica", case=False, na=False)
    ].shape[0]

    st.divider()

    c1, c2, c3, c4 = st.columns(4)

    c1.metric("🚨 Incidencias activas", total)
    c2.metric("🟡 Abiertas", abiertas)
    c3.metric("🔵 En proceso", proceso)
    c4.metric("🔴 Críticas", criticas)

    st.divider()

    st.subheader("🚦 Distribución por estatus y prioridad")

    pintar_matriz_incidencias(df_filtrado)

    st.divider()

    st.subheader("📋 Detalle de incidencias por transporte")

    columnas_detalle = [
        "folio_incidencia",
        "fecha",
        "estatus",
        "prioridad",
        "tipo_incidencia",
        "transporte_mostrar",
        "transportista",
        "vehiculo",
        "placas",
        "operador",
        "responsable",
        "cliente",
        "destino",
        "descripcion_incidencia",
        "causa",
        "solucion",
        "observaciones"
    ]

    columnas_existentes = [
        col for col in columnas_detalle
        if col in df_filtrado.columns
    ]

    st.dataframe(
        df_filtrado[columnas_existentes],
        use_container_width=True,
        height=350,
        hide_index=True
    )

    st.divider()

    excel_dashboard = generar_excel_incidencias(
        df_filtrado[columnas_existentes],
        total,
        abiertas,
        proceso,
        criticas
    )

    st.download_button(
        label="📥 Exportar dashboard Excel",
        data=excel_dashboard,
        file_name=(
            f"dashboard_incidencias_transportes_"
            f"{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        ),
        mime=(
            "application/vnd.openxmlformats-officedocument."
            "spreadsheetml.sheet"
        ),
        use_container_width=True
    )

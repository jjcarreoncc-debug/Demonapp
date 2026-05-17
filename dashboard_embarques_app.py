import streamlit as st
import pandas as pd
import sqlite3
import altair as alt
from io import BytesIO
from datetime import datetime

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows

from sigem_db import get_db_path


# =====================================================
# OBTENER EMBARQUES ACTIVOS
# =====================================================

def obtener_embarques():

    conn = sqlite3.connect(
        get_db_path("logistica")
    )

    query = """
        SELECT
            e.folio_embarque,
            e.folio_hoja_carga,
            e.folio_ruta,
            e.codigo_transporte,
            e.pedido,
            e.fecha,
            e.cliente,
            e.destino,
            e.transportista,
            e.vehiculo,
            e.placas,
            e.operador,
            e.ruta,
            e.estatus,

            i.tipo_incidencia,
            i.prioridad AS prioridad_incidencia,
            i.estatus AS estatus_incidencia,
            i.descripcion_incidencia,
            i.fecha AS fecha_incidencia

        FROM embarques e

        LEFT JOIN (

            SELECT
                i1.*

            FROM incidencias i1

            INNER JOIN (

                SELECT
                    folio_embarque,
                    MAX(fecha) AS max_fecha

                FROM incidencias

                GROUP BY folio_embarque

            ) ult

            ON i1.folio_embarque = ult.folio_embarque
            AND i1.fecha = ult.max_fecha

        ) i

        ON e.folio_embarque = i.folio_embarque

        WHERE IFNULL(e.estatus, '') IN (
            'Pendiente carga',
            'En almacén',
            'En patio',
            'Ya salió',
            'En tránsito',
            'Cargado'
        )

        ORDER BY e.fecha DESC,
                 e.folio_embarque DESC
    """

    df = pd.read_sql_query(query, conn)

    conn.close()

    return df


# =====================================================
# PREPARAR TRANSPORTES
# =====================================================

def preparar_transportes(df):

    df = df.copy()

    for col in [
        "transportista",
        "vehiculo",
        "placas",
        "operador",
        "ruta",
        "cliente",
        "destino",
        "estatus"
    ]:

        if col not in df.columns:
            df[col] = ""

        df[col] = (
            df[col]
            .fillna("")
            .astype(str)
        )

    df["transporte_display"] = (

        df["transportista"]
        .replace("", "SIN TRANSPORTISTA")

        + " - " +

        df["vehiculo"]
        .replace("", "SIN VEHÍCULO")

        + " - " +

        df["placas"]
        .replace("", "SIN PLACAS")
    )

    return df


# =====================================================
# RESUMEN TRANSPORTES
# =====================================================

def generar_df_transportes(df):

    resumen = (

        df.groupby(
            [
                "transporte_display",
                "transportista",
                "vehiculo",
                "placas",
                "operador",
                "ruta",
                "estatus"
            ],
            dropna=False
        )

        .agg(
            embarques=("folio_embarque", "nunique"),
            hojas_carga=("folio_hoja_carga", "nunique"),
            clientes=("cliente", "nunique"),
            destinos=("destino", "nunique")
        )

        .reset_index()
    )

    return resumen


# =====================================================
# EXPORTAR EXCEL
# =====================================================

def generar_excel_dashboard(
    df_transportes,
    df_detalle,
    total_transportes,
    total_embarques,
    pendientes,
    transito
):

    wb = Workbook()

    ws = wb.active

    ws.title = "Dashboard transportes"

    azul = PatternFill(
        start_color="1F4E78",
        end_color="1F4E78",
        fill_type="solid"
    )

    borde = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin")
    )

    font_blanca = Font(
        color="FFFFFF",
        bold=True
    )

    ws["A1"] = "SIGEM - Dashboard transportes activos"

    ws["A1"].font = Font(
        bold=True,
        size=16
    )

    ws["A3"] = "Transportes activos"
    ws["B3"] = total_transportes

    ws["A4"] = "Embarques activos"
    ws["B4"] = total_embarques

    ws["A5"] = "Pendiente carga"
    ws["B5"] = pendientes

    ws["A6"] = "En tránsito"
    ws["B6"] = transito

    for celda in [
        "A3",
        "A4",
        "A5",
        "A6"
    ]:

        ws[celda].fill = azul
        ws[celda].font = font_blanca
        ws[celda].border = borde

    ws2 = wb.create_sheet(
        title="Transportes"
    )

    for r in dataframe_to_rows(
        df_transportes,
        index=False,
        header=True
    ):

        ws2.append(r)

    for cell in ws2[1]:

        cell.fill = azul
        cell.font = font_blanca
        cell.border = borde

    ws3 = wb.create_sheet(
        title="Detalle embarques"
    )

    for r in dataframe_to_rows(
        df_detalle,
        index=False,
        header=True
    ):

        ws3.append(r)

    for cell in ws3[1]:

        cell.fill = azul
        cell.font = font_blanca
        cell.border = borde

    buffer = BytesIO()

    wb.save(buffer)

    buffer.seek(0)

    return buffer


# =====================================================
# MATRIZ ESTATUS
# =====================================================

def pintar_matriz_estatus(df_transportes):

    orden_estatus = [
        "Pendiente carga",
        "En almacén",
        "En patio",
        "Cargado",
        "Ya salió",
        "En tránsito"
    ]

    mapa_colores = {

        "Pendiente carga": "#6B7280",
        "En almacén": "#7C3AED",
        "En patio": "#F97316",
        "Cargado": "#0EA5E9",
        "Ya salió": "#EAB308",
        "En tránsito": "#2563EB"
    }

    df_grafica = df_transportes.copy()

    ancho_grafica = max(
        1300,
        len(df_grafica["transporte_display"].unique()) * 180
    )

    chart = (

        alt.Chart(df_grafica)

        .mark_rect(
            cornerRadius=5,
            width=90,
            height=28
        )

        .encode(

            x=alt.X(
                "transporte_display:N",
                title="Transporte",
                sort=None,
                axis=alt.Axis(
                    labelAngle=-45,
                    labelLimit=220,
                    grid=True,
                    tickSize=0
                )
            ),

            y=alt.Y(
                "estatus:N",
                sort=orden_estatus,
                title="Estatus",
                axis=alt.Axis(
                    grid=True,
                    tickSize=0
                )
            ),

            color=alt.Color(
                "estatus:N",
                scale=alt.Scale(
                    domain=list(
                        mapa_colores.keys()
                    ),
                    range=list(
                        mapa_colores.values()
                    )
                ),
                legend=None
            ),

            tooltip=[

                alt.Tooltip(
                    "transporte_display:N",
                    title="Transporte"
                ),

                alt.Tooltip(
                    "transportista:N",
                    title="Transportista"
                ),

                alt.Tooltip(
                    "vehiculo:N",
                    title="Vehículo"
                ),

                alt.Tooltip(
                    "placas:N",
                    title="Placas"
                ),

                alt.Tooltip(
                    "operador:N",
                    title="Operador"
                ),

                alt.Tooltip(
                    "ruta:N",
                    title="Ruta"
                ),

                alt.Tooltip(
                    "estatus:N",
                    title="Estatus"
                ),

                alt.Tooltip(
                    "embarques:Q",
                    title="Embarques"
                ),

                alt.Tooltip(
                    "hojas_carga:Q",
                    title="Hojas carga"
                )
            ]

        )

        .properties(
            width=ancho_grafica,
            height=480
        )

        .configure_axis(
            grid=True,
            gridColor="#D1D5DB",
            gridDash=[3, 2],
            domain=False,
            labelColor="#374151",
            titleColor="#111827"
        )

        .configure_view(
            stroke="#D1D5DB"
        )

    )

    st.altair_chart(
        chart,
        use_container_width=False
    )


# =====================================================
# APP
# =====================================================

def dashboard_embarques_app():

    st.title(
        "🚛 Dashboard transportes activos"
    )

    st.caption(
        "Trazabilidad operativa de transportes y embarques activos."
    )

    try:

        df = obtener_embarques()

    except Exception as e:

        st.error(
            "❌ Error consultando embarques activos."
        )

        st.exception(e)

        return

    if df.empty:

        st.warning(
            "No existen transportes con embarques activos."
        )

        return

    df = preparar_transportes(df)

    df_filtrado = df.copy()

    df_transportes = generar_df_transportes(
        df_filtrado
    )

    total_transportes = (
        df_filtrado[
            "transporte_display"
        ].nunique()
    )

    total_embarques = (
        df_filtrado[
            "folio_embarque"
        ].nunique()
    )

    pendientes = df_filtrado[
        df_filtrado["estatus"]
        .astype(str)
        .str.contains(
            "Pendiente carga",
            case=False,
            na=False
        )
    ].shape[0]

    transito = df_filtrado[
        df_filtrado["estatus"]
        .astype(str)
        .str.contains(
            "tránsito|transito",
            case=False,
            na=False
        )
    ].shape[0]

    st.divider()

    c1, c2, c3, c4 = st.columns(4)

    c1.metric(
        "🚛 Transportes activos",
        total_transportes
    )

    c2.metric(
        "📦 Embarques activos",
        total_embarques
    )

    c3.metric(
        "⏳ Pendiente carga",
        pendientes
    )

    c4.metric(
        "🚚 En tránsito",
        transito
    )

    st.divider()

    st.subheader(
        "🚦 Trazabilidad por transporte"
    )

    pintar_matriz_estatus(
        df_transportes
    )

    st.divider()

    st.subheader(
        "🚛 Transportes activos"
    )

    st.dataframe(
        df_transportes,
        use_container_width=True,
        hide_index=True,
        height=280
    )

    st.divider()

    st.subheader(
        "📦 Detalle embarques asociados"
    )

    columnas_detalle = [

        "folio_embarque",

        "folio_hoja_carga",

        "pedido",

        "fecha",

        "cliente",

        "destino",

        "transportista",

        "vehiculo",

        "placas",

        "operador",

        "ruta",

        "estatus",

        "tipo_incidencia",

        "prioridad_incidencia",

        "estatus_incidencia"
    ]

    columnas_detalle = [
        col
        for col in columnas_detalle
        if col in df_filtrado.columns
    ]

    st.dataframe(
        df_filtrado[columnas_detalle],
        use_container_width=True,
        hide_index=True,
        height=420
    )

    st.divider()

    excel_dashboard = generar_excel_dashboard(
        df_transportes,
        df_filtrado[columnas_detalle],
        total_transportes,
        total_embarques,
        pendientes,
        transito
    )

    st.download_button(
        label="📥 Exportar dashboard Excel",
        data=excel_dashboard,
        file_name=(
            f"dashboard_transportes_activos_"
            f"{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        ),
        mime=(
            "application/vnd.openxmlformats-officedocument."
            "spreadsheetml.sheet"
        ),
        use_container_width=True
    )


if __name__ == "__main__":

    dashboard_embarques_app()
